# 尚米 — Excel 导出模块
# 把商家数据导出为 .xlsx 表格

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


def export_to_excel(businesses, filepath, keyword=""):
    """
    导出商家数据到 Excel 文件

    参数:
        businesses: 商家数据列表（字典）
        filepath: 保存路径（如 /tmp/尚米_北京川菜馆.xlsx）
        keyword: 搜索关键词（用于文件名中的描述）

    返回:
        filepath（成功时）或 None（失败时）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商家数据"

    # ===== 样式定义 =====
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    phone_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")  # 电话蓝字带下划线

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ===== 表头 =====
    headers = ["序号", "商家名称", "电话", "地址", "城市", "区/县", "分类", "评分", "人均消费/价格", "营业时间", "来源"]
    col_widths = [8, 30, 18, 40, 10, 12, 20, 8, 12, 22, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ===== 数据行 =====
    for row_idx, biz in enumerate(businesses, 2):
        row_data = [
            row_idx - 1,                                 # 序号
            biz.get("name", ""),                          # 商家名称
            biz.get("phone", ""),                         # 电话
            biz.get("address", ""),                       # 地址
            biz.get("city", ""),                          # 城市
            biz.get("district", ""),                      # 区/县
            biz.get("category", ""),                      # 分类
            biz.get("rating", ""),                        # 评分
            biz.get("cost", "") or biz.get("price", ""),  # 人均消费/价格
            biz.get("shop_hours", ""),                    # 营业时间（百度独有）
            biz.get("source", ""),                        # 来源
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

        # 奇偶行交替底色（方便阅读）
        if row_idx % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # ===== 冻结第一行（表头始终可见） =====
    ws.freeze_panes = "A2"

    # ===== 自动筛选 =====
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(businesses) + 1}"

    # ===== 保存 =====
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath
